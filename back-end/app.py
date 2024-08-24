import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Carrega a planilha de clientes
planilha_cliente = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_cliente['Sheet1']

# Cria uma nova planilha de fechamento se não existir
try:
    planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
except FileNotFoundError:
    planilha_fechamento = openpyxl.Workbook()
    planilha_fechamento.save('planilha_fechamento.xlsx')

pagina_fechamento = planilha_fechamento['Sheet1']

# Inicializa o driver do Selenium
driver = webdriver.Chrome()

# Abre a página do site
driver.get('http://localhost:8000')

# Itera sobre as linhas da planilha, começando da segunda linha
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    # Aguarda um pouco antes de interagir com a página
    sleep(5)
    
    # Encontra o campo de pesquisa de CPF, limpa e insere o novo CPF
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    
    # Encontra e clica no botão de consultar
    botao_consultar = driver.find_element(By.XPATH, "//button[@class='btn-consultar']")
    botao_consultar.click()
    sleep(4)
    
    # Encontra o status de pagamento
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    
    # Adiciona os dados à planilha de fechamento
    data_pagamento_limpo = ''
    metodo_pagamento_limpo = ''
    
    if status.text == 'em dia':
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

    # Adiciona os dados à planilha de fechamento
    linha_fechamento = [nome, valor, cpf, vencimento, status.text, data_pagamento_limpo, metodo_pagamento_limpo]
    pagina_fechamento.append(linha_fechamento)

    # Salva a planilha de fechamento após cada consulta
    planilha_fechamento.save('planilha_fechamento.xlsx')

# Fecha o driver do Selenium
driver.quit()
